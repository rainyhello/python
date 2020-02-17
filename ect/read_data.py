"""
Created on Mon Feb 17 10:46:42 2020

@author: Tsinghua
"""

#import numpy as np
import pandas as pd
import os
from openpyxl import load_workbook

#具体文件的结构可以进入E盘或者D盘的log目录下查看清楚
#windows的文件或者文件夹路径用的是反斜杠，而python中要用斜杠或者双反斜杠
#因为反斜杠是转义字符，比如//的意思是/，还有很多其他转义字符，请百度
#特殊的反斜杠放在一行代码最后面，则表示拆分成多行，连接下一行的意思
dir_root = 'E:/log'
#dir_a = 'E:/log/A190820C31N81'
dir_ch = 'E:\\log\\A190820C31N81\\04  81fF'
#dir_txt = 'E:\\log\\A190820C31N81\\04  81fF\\CH1-CH2'
#dir_picture = 'E:\log\picture'
dir_excel = 'E:/log/data_process.xlsx'
#python中一行较长代码分成多行代码，需要在每行最后添加反斜杠连接，如下所示
#dir_data = 'E:/log/A190820C31N81/04  81fF/CH1-CH2\
#            /ECT1Data_RAW_2020-02-15_12-55-43.txt2'

cc=['21fF','29fF','51fF','81fF','150fF','297fF','394fF','583fF',
    '708fF','1057fF','1430fF']
#bb=np.array([22,29,51,81,150,297,394,583,708,1057,1430])
#下面是遍历某个文件夹下面的所有文件和目录的一个模块
def all_txt(folder):
    import os
    for root, dirs, files in os.walk(folder):
        for file in files:
            # 获取文件路径
            yield (os.path.join(root, file))
 
#对CH1-CH2到CH7-CH8的文件夹进行编号排序
u = os.listdir(dir_ch)
dic_ch = {}
n = 0
for i in u:
    if 'CH' in i:
        dic_ch[i] = n
        n=n+1

#对电容大小进行排序
#v = os.listdir(dir_a)
#dic_a = {}
#n = 0
#for i in v:
#    if 'fF' in i:
#        dic_a[i] = n
#        n = n + 1

#对所有的板卡进行编号和排序
u = os.listdir(dir_root)
dic_root = {}
n = 0
for i in u:
    if 'A19' in i:
        dic_root[i] = n
        n = n + 1

#打开Excel文件
wb = load_workbook(dir_excel)#生成一个已存在的wookbook对象
wb1 = wb.active#激活sheet

for i in all_txt(dir_root):
    if '.txt' in i and 'ECT' in i:#所有txt文件名字都是ECT+时间命名
#       print(i)
        data = pd.read_table(i,header=None)#读取txt文件数据
        d = i.split('\\')[3]#第四级目录-28通道
        c = i.split('\\')[2][:2]#第三级目录的前两个字符-电容序号
        b = i.split('\\')[1]#第二级目录-板卡
        a = i.split('\\')[0]#第一级目录-根目录
        x = dic_ch[d]#0-27通道
        y = dic_root[b]#板卡编号0-n
        z = int(c)#电容编号0-11,而文件夹命名是从1开始
        res = data.values[:,x+1].mean()#data第一列是时间
        
        wb1.cell(5,x+3,d)#填写28通道编号在第五行的3-31列
        wb1.cell(y*11+z+5,1,b)#填写板卡序列号在第一列的某行至某行
        wb1.cell(y*11+z+5,2,cc[z-1])#填写电容序号
        wb1.cell(5,1,'板卡序列号')
        wb1.cell(5,2,'被测电容值')
        
        wb1.cell(y*11+z+5,x+3,res)#往sheet中的第t*5+1行第n+2列写入均值
            
wb.save(dir_excel)#保存
wb.close()
